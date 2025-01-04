'''
Author: HDJ
StartDate: 2024-05-15 00:00:00
LastEditTime: 2024-12-28 22:03:58
FilePath: \pythond:\LocalUsers\Goodnameisfordoggy-Gitee\JD-Automated-Tools\JD-PersDataExporter\src\storage\dataStorageToExcel.py
Description: 

                *       写字楼里写字间，写字间里程序员；
                *       程序人员写程序，又拿程序换酒钱。
                *       酒醒只在网上坐，酒醉还来网下眠；
                *       酒醉酒醒日复日，网上网下年复年。
                *       但愿老死电脑间，不愿鞠躬老板前；
                *       奔驰宝马贵者趣，公交自行程序员。
                *       别人笑我忒疯癫，我笑自己命太贱；
                *       不见满街漂亮妹，哪个归得程序员？    
Copyright (c) 2024-2025 by HDJ, All Rights Reserved. 
'''
import os
import string
import zipfile
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment

from src.dataPortector import OrderExportConfig
from src.logger import get_logger
LOG = get_logger()


class ExcelStorage:
    def __init__(self, data: list[dict]= [], header_needed: list = [], file_name: str = "JD_order_info", sheet_name: str = "Sheet1"):
        self.__data = data  # 需python3.7及以上，利用字典键值对的插入顺序。
        self.__header_needed = header_needed  # 用户选择的表头字段列表
        # 获取配置文件
        self.__config = OrderExportConfig().load_from_json()

        if any(file_name.endswith(ext) for ext in ['.xlsx', '.xlsm', '.xltx', '.xltm']):
            self.__file_name = file_name
        else:
            self.__file_name = file_name + '.xlsx'# 设置生成文件的文件名
        self.__sheet_name = sheet_name # 设置生成的工作簿名
        self.__existent_order_id = None # 现有表中的id
        self.__output_fields = self.__define_output_fields()
    
    def __define_output_fields(self):
        """ 
        定义输出字段组合：

        筛除未定义的字段名。

        Return: (list) 处理后的字段名称列表。
        """
        field_items = [value for _, value in self.__config.excel_storage_settings["headers_settings"].items()]
        # LOG.debug(f'field_items: {field_items}\n')

        all_fields_name = [field.get('name', '') for field in field_items] # 全部已定义字段的名称
        # LOG.debug(f'all_fields_name: {all_fields_name}\n')

        # 保证表中有 "订单编号"，作为订单唯一性标志
        if '订单编号' not in self.__header_needed:
            self.__header_needed = ['订单编号'] + self.__header_needed
        # 筛除未定义的字段名
        self.__header_needed = [item for item in self.__header_needed if item in all_fields_name]
        # LOG.debug(f'__header_needed: {self.__header_needed}\n')
        
        # 收集不需要的字段名
        # fields_not_needed = [item for item in all_field if item not in self.__header_needed]

        field_items_sorted = []
        # 重新排序field_items
        for name in self.__header_needed:
            for item in field_items:
                if item.get('name', '') == name:
                    field_items_sorted.append(item)
        # LOG.debug(f'field_items_sorted: {field_items_sorted}\n')
        return [item.get('name', '') for item in field_items_sorted]
    
    def __is_file_exists(self):
        if os.path.exists(self.__file_name):
            LOG.info(f"Reading file: {self.__file_name}, sheet: {self.__sheet_name}")
            return True
        else:
            LOG.info("文件不存在, 将新建Excel文件。")

    def __is_file_locked(self):
        """检测文件是否被占用"""
        try:
            # 尝试以写入模式打开文件
            with open(self.__file_name, 'a') as f:
                pass  # 如果文件可以被打开，则直接跳过
            return False  # 文件可以被打开，未被占用
        except PermissionError as e:
            LOG.error(f"Permission error: {e}. 请确保文件没有在其他程序中打开，并且您有写入权限。")
            return True  # 文件被占用
    
    def __create_new_file(self):
        """ 创建新的Excel文件 """
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = self.__sheet_name
        
        # 写入列名
        sheet.append(self.__output_fields)
        # 保存工作簿
        workbook.save(self.__file_name)
        LOG.info(f"新文件 '{self.__file_name}' 创建成功。")
    
    def __append_to_excel(self, df):
        """ 在现有的 Excel 文件末尾添加新内容 """
        try:
            # 尝试打开现有的工作簿
            workbook = load_workbook(self.__file_name)
            sheet = workbook[self.__sheet_name]
    
            # 将 DataFrame 的每一行添加到工作表
            for row in df.itertuples(index=False, name=None):
                sheet.append(row)
                
            # 保存工作簿
            workbook.save(self.__file_name)
            LOG.info(f"数据成功添加到 '{self.__file_name}'，新增数据{len(self.__data)}条")
        except FileNotFoundError:
            pass
        except PermissionError as e:
            LOG.error(f"Permission error: {e}. 请确保文件没有在其他程序中打开，并且您有写权限。")
        except Exception as e:
            LOG.critical(f"添加到Excel时发生错误: {e}")

    def __adjust_column_format(self):
        """ 调整 Excel 表头格式，以及单元格数据类型 """
        try:
            workbook = load_workbook(self.__file_name)
            worksheet = workbook.active

            uppercase_letters = [letter for letter in string.ascii_uppercase]  # 大写字母A~Z，用于表头设置
            index = 0
            default_width = 16
            for header in self.__output_fields:  # 获取用户选择的表头名称，以此为顺序
                for key, header_item in self.__config.excel_storage_settings["headers_settings"].items():  # 获取表头的设置信息
                    if header_item["name"] == header:
                        # 设置表头宽度
                        worksheet.column_dimensions[uppercase_letters[index]].width = float(header_item.get("width", default_width))
                        index += 1  # 下一个表头字母的索引
                        break
                    
                    match header:
                        case "实付金额":
                            actual_payment_amount_col = self.__output_fields.index("实付金额") + 1
                            # 设置 amount 列的所有单元格为人民币格式
                            for row in worksheet.iter_rows(min_col=actual_payment_amount_col, max_col=actual_payment_amount_col, min_row=2, max_row=worksheet.max_row):
                                for cell in row:
                                    # 确保数据为数字类型
                                    if isinstance(cell.value, (int, float)):
                                        try:
                                            cell.value = float(cell.value) 
                                        except ValueError:
                                            # 如果无法转换为浮点数，可能数据格式有问题，跳过处理
                                            continue
                                    cell.number_format = '¥#,##0.00' # 设置单元格格式为人民币 ¥
                                    cell.alignment = Alignment(horizontal="right") # 设置单元格内容右对齐
                        case "商品总价":
                            product_total_price_col = self.__output_fields.index("商品总价") + 1
                            # 设置 amount 列的所有单元格为人民币格式
                            for row in worksheet.iter_rows(min_col=product_total_price_col, max_col=product_total_price_col, min_row=2, max_row=worksheet.max_row):
                                for cell in row:
                                    # 确保数据为数字类型
                                    if isinstance(cell.value, (int, float)):
                                        try:
                                            cell.value = float(cell.value) 
                                        except ValueError:
                                            # 如果无法转换为浮点数，可能数据格式有问题，跳过处理
                                            continue
                                    cell.number_format = '¥#,##0.00' # 设置单元格格式为人民币 ¥
                                    cell.alignment = Alignment(horizontal="right") # 设置单元格内容右对齐
                        case "订单返豆":
                            jingdou_increment_col = self.__output_fields.index('订单返豆') + 1
                            for row in worksheet.iter_rows(min_col=jingdou_increment_col, max_col=jingdou_increment_col, min_row=2, max_row=worksheet.max_row):
                                for cell in row:
                                    # 确保数据为数字类型
                                    if isinstance(cell.value, int):
                                        cell.number_format = '0'  # 设置单元格格式为整数
                        case "订单用豆":
                            jingdou_decrement_col = self.__output_fields.index('订单用豆') + 1
                            for row in worksheet.iter_rows(min_col=jingdou_decrement_col, max_col=jingdou_decrement_col, min_row=2, max_row=worksheet.max_row):
                                for cell in row:
                                    # 确保数据为数字类型
                                    if isinstance(cell.value, int):
                                        cell.number_format = '0'  # 设置单元格格式为整数
                        case "下单时间":
                            order_time_col = self.__output_fields.index('下单时间') + 1
                            for row in worksheet.iter_rows(min_col=order_time_col, max_col=order_time_col, min_row=2, max_row=worksheet.max_row):
                                for cell in row:
                                    cell.number_format = "yyyy-mm-dd hh:mm:ss" # 设置单元格的日期时间格式
                                    cell.alignment = Alignment(horizontal="left")
                # 表头设置水平和垂直居中对齐
                for cell in worksheet[1]:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                # 保存修改后的 Excel 文件
                workbook.save(self.__file_name)
        except zipfile.BadZipFile:
            pass
        except Exception as e:
            LOG.critical(e)
            
    def save(self):
        """ 
        数据储存 
        """
        if not self.__is_file_exists():
            self.__create_new_file()
        else:
            if not self.__is_file_locked():
                # 读取Excel文件
                df = pd.read_excel(self.__file_name, sheet_name=self.__sheet_name, engine='openpyxl')
                # 获取表中 "订单编号" 字段下的所有值
                df['订单编号'] = df['订单编号'].fillna(0) # 将 NaN 值替换为一个默认值（例如0或其他适合的值）
                df['订单编号'] = df['订单编号'].astype(int) # 将列转换为整数
                self.__existent_order_id = df['订单编号'].tolist()
                self.__existent_order_id = [str(item) for item in self.__existent_order_id]
                self.__data = [order for order in self.__data if order.get('订单编号') not in self.__existent_order_id]

        try:
            if self.__data:
                if not self.__is_file_locked():
                    # 创建DataFrame
                    df = pd.DataFrame(self.__data, columns=self.__output_fields)
                    self.__append_to_excel(df)
                    # 设置Excel文件列宽
                    self.__adjust_column_format()
            else:
                LOG.info(f'没有新数据, 储存结束。')
                return
        except zipfile.BadZipFile:
            LOG.error(f"BadZipFile error: The file is not a valid zip file (Excel file): {self.__file_name}")
    
    def get_all_sheets_name(self):
        """获取Excel文件中的所有数据表名称"""
        if self.__is_file_exists():
            if not self.__is_file_locked():
                # 读取Excel文件
                xls = pd.ExcelFile(self.__file_name, engine='openpyxl')
                # 获取所有 Sheet 名称
                sheet_names = xls.sheet_names
                return sheet_names
