'''
Author: HDJ
StartDate: 2024-05-15 00:00:00
LastEditTime: 2024-06-03 13:28:42
FilePath: \pythond:\LocalUsers\Goodnameisfordoggy-Gitee\jd-pers-data-exporter\src\storage\dataStorageToExcel.py
Description: 

                *       写字楼里写字间，写字间里程序员；
                *       程序人员写程序，又拿程序换酒钱。
                *       酒醒只在网上坐，酒醉还来网下眠；
                *       酒醉酒醒日复日，网上网下年复年。
                *       但愿老死电脑间，不愿鞠躬老板前；
                *       奔驰宝马贵者趣，公交自行程序员。
                *       别人笑我忒疯癫，我笑自己命太贱；
                *       不见满街漂亮妹，哪个归得程序员？    
Copyright (c) 2024 by HDJ, All Rights Reserved. 
'''
import os
import string
import pandas as pd
import zipfile
from openpyxl import load_workbook, Workbook

from ..dataPortector import ConfigManager


class ExcelStorage:
    def __init__(self, data: list[dict], header_needed: list, file_name: str):
        self.__data = data  # 需python3.7及以上，利用字典键值对的插入顺序。
        self.__header_needed = header_needed  # 用户选择的表头字段列表
        # 获取配置文件
        self.__configManager = ConfigManager()
        self.__config = self.__configManager.get_excel_config()
        # 设置生成文件的文件名
        self.__file_name = file_name
        self.__sheet_name = 'Sheet1'
        self.__existent_order_id = None
    
    def is_file_exists(self):
        if os.path.exists(self.__file_name):
            print(f"Reading file: {self.__file_name}, sheet: {self.__sheet_name}")
            return True
        else:
            print("文件不存在, 将新建Excel文件。")

    def is_file_locked(self):
        """检测文件是否被占用"""
        try:
            # 尝试以写入模式打开文件
            with open(self.__file_name, 'a') as f:
                pass  # 如果文件可以被打开，则直接跳过
            return False  # 文件可以被打开，未被占用
        except PermissionError as e:
            print(f"Permission error: {e}. 请确保文件没有在其他程序中打开，并且您有写权限。")
            return True  # 文件被占用
    
    def create_new_file(self):
        """ 创建新的Excel文件 """
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = self.__sheet_name
        
        # 写入列名
        sheet.append(self.__header_needed)
        # 保存工作簿
        workbook.save(self.__file_name)
        print(f"新文件 '{self.__file_name}' 创建成功。")
    
    def append_to_excel(self, df):
        """ 在现有的 Excel 文件末尾添加新内容 """
        try:
           # 尝试打开现有的工作簿
            workbook = load_workbook(self.__file_name)
            sheet = workbook[self.__sheet_name]
    
            # 将 DataFrame 的每一行添加到工作表
            for row in df.itertuples(index=False, name=None):
                sheet.append(row)
                
            # 保存工作簿
            workbook.save(self.__file_name)
            print(f"数据成功添加到 {self.__file_name}")
        except FileNotFoundError:
            pass
        except PermissionError as e:
            print(f"Permission error: {e}. 请确保文件没有在其他程序中打开，并且您有写权限。")
        except Exception as e:
            print(f"添加到Excel时发生错误: {e}")
       
    def adjust_column_width(self):
        """ 调整 Excel 表头宽度 """
        try:
            workbook = load_workbook(self.__file_name)
            worksheet = workbook.active

            uppercase_letters = [letter for letter in string.ascii_uppercase]  # 大写字母A~Z，用于表头设置
            index = 0
            default_width = 16
            for header in self.__header_needed:  # 获取用户选择的表头名称
                for header_item in self.__config['header_items']:  # 获取该表头的信息
                    if header_item['name'] == header:
                        # 设置表头宽度
                        worksheet.column_dimensions[uppercase_letters[index]].width = float(header_item.get('width', default_width))
                        index += 1  # 下一个表头字母的索引
                        break
            # 保存修改后的 Excel 文件
            workbook.save(self.__file_name)
        except zipfile.BadZipFile:
            pass
        except Exception as e:
            print(f"An error occurred: {e}")
            
    def save(self):
        """ 
        数据储存 
        """
        if not self.is_file_exists():
            self.create_new_file()
        else:
            # 读取Excel文件
            df = pd.read_excel(self.__file_name, sheet_name=self.__sheet_name, engine='openpyxl')
            # 获取表中order_id字段下的所有值
            self.__existent_order_id = df['order_id'].tolist()
            self.__data = [order for order in self.__data if int(order.get('order_id')) not in self.__existent_order_id]
            # print(self.__existent_order_id)
            # print(self.__data)

        try:
            if self.__data:
                if not self.is_file_locked():
                    # 创建DataFrame
                    df = pd.DataFrame(self.__data, columns=self.__header_needed)
                    self.append_to_excel(df)
                    # 设置Excel文件列宽
                    self.adjust_column_width()
            else:
                print(f'没有新数据, 储存结束。')
                return
        except zipfile.BadZipFile:
            print(f"BadZipFile error: The file is not a valid zip file (Excel file): {self.__file_name}")
        
    